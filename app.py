"""
=============================================================================
app.py  --  INTERFEJS GRAFICZNY
=============================================================================

URUCHOMIENIE

    conda activate sirna
    cd sciezka/do/sirna_pipeline
    streamlit run app.py

Autor: Antonina Jarecka
=============================================================================
"""

import io
import re
import tempfile

import streamlit as st

import hosts
import design
import scoring
import constructs

try:
    import offtarget
    OFFTARGET_OK = True
except ImportError:
    OFFTARGET_OK = False

try:
    import conservation
    KONSERWATYWNOSC_OK = True
except ImportError:
    KONSERWATYWNOSC_OK = False

try:
    import oligos
    OLIGOS_OK = True
except ImportError:
    OLIGOS_OK = False

try:
    import ncbi
    NCBI_OK = True
except ImportError:
    NCBI_OK = False

try:
    import report
    REPORT_OK = True
except ImportError:
    REPORT_OK = False


st.set_page_config(page_title='siRNA Design Studio', page_icon='🧬',
                   layout='wide', initial_sidebar_state='expanded')


# ============================================================================
# STYL
# ============================================================================

STYL = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Spectral:wght@500;600;700&family=IBM+Plex+Sans:wght@400;450;500;600;700&family=IBM+Plex+Mono:wght@500;600&display=swap');

:root {
  --fiolet:        #6B4E96;
  --fiolet-ciemny: #4E3670;
  --fiolet-jasny:  #9B7BC8;
  --roz:           #D4739E;
  --roz-jasny:     #F0C4D9;
  --zloto:         #B8963F;
  --zloto-jasne:   #E3CE93;
  --srebro:        #98A0AC;
  --srebro-jasne:  #E4E8EE;
  --tekst:         #2B2435;
  --tekst-mniej:   #5F5670;
}

/* ---------- BACKGROUND: soft hexagonal lattice over a colour gradient ---- */
.stApp {
  background:
    radial-gradient(ellipse at 6% 0%,    rgba(155,123,200,0.22) 0%, transparent 46%),
    radial-gradient(ellipse at 96% 4%,   rgba(212,115,158,0.19) 0%, transparent 44%),
    radial-gradient(ellipse at 78% 62%,  rgba(184,150,63,0.13)  0%, transparent 42%),
    radial-gradient(ellipse at 14% 88%,  rgba(152,160,172,0.16) 0%, transparent 44%),
    linear-gradient(158deg, #FCFAFD 0%, #F6F1F8 32%, #F9F2F5 62%, #FAF6F0 100%);
  background-attachment: fixed;
}

/* A single quiet motif: hexagonal lattice, as in a chemical skeleton.
   Kept at very low opacity so it reads as texture, not decoration. */
.stApp::before {
  content: ''; position: fixed; inset: 0; pointer-events: none; z-index: 0;
  opacity: 0.055;
  background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='112' height='128' viewBox='0 0 112 128'%3E%3Cg fill='none' stroke='%236B4E96' stroke-width='1.15'%3E%3Cpath d='M56 3 L104 30 L104 84 L56 111 L8 84 L8 30 Z'/%3E%3Cpath d='M56 67 L104 94 L104 128'/%3E%3Cpath d='M56 67 L8 94 L8 128'/%3E%3Cpath d='M56 -61 L104 -34 L104 30'/%3E%3C/g%3E%3C/svg%3E");
  background-repeat: repeat;
  background-size: 112px 128px;
}

/* ---------- hide the Streamlit toolbar and deploy button ---------- */
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="stStatusWidget"],
.stAppDeployButton,
#MainMenu,
footer { display: none !important; visibility: hidden !important; }
header[data-testid="stHeader"] { background: transparent !important;
  height: 0 !important; }

.main .block-container {
  position: relative; z-index: 1;
  padding-top: 2rem; max-width: 1340px;
}

/* ---------- TYPOGRAFIA (pogrubiona) ---------- */
html, body, [class*="css"], .stMarkdown, p, li, label, span, div {
  font-family: 'IBM Plex Sans', 'Segoe UI', -apple-system, sans-serif;
  color: var(--tekst);
  font-weight: 500;
}
.stMarkdown p, .stMarkdown li { font-size: 1.0rem; line-height: 1.62; font-weight: 450; }

h1 {
  font-family: 'Spectral', 'Palatino Linotype', Georgia, serif !important;
  font-weight: 600 !important;
  font-size: 2.5rem !important;
  letter-spacing: -0.022em;
  background: linear-gradient(102deg, var(--fiolet-ciemny) 0%, var(--fiolet) 32%, var(--roz) 66%, var(--zloto) 100%);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text;
  margin-bottom: 0.05rem !important;
}
h2 { font-family:'Spectral','Palatino Linotype',Georgia,serif !important; font-weight:700 !important;
     color: var(--fiolet-ciemny) !important; font-size: 1.55rem !important;
     letter-spacing: -0.012em; }
h3 { font-family:'Spectral','Palatino Linotype',Georgia,serif !important; font-weight:700 !important;
     color: var(--fiolet) !important; font-size: 1.3rem !important;
     letter-spacing: -0.008em; }
h4 { font-family:'IBM Plex Sans','Segoe UI',sans-serif !important; font-weight:700 !important;
     color: var(--fiolet) !important; font-size: 1.04rem !important;
     letter-spacing: 0.012em; }

.stCaption, [data-testid="stCaptionContainer"] {
  color: var(--tekst-mniej) !important; font-weight: 450 !important; font-size: 0.87rem !important;
}

/* ---------- PASEK BOCZNY ---------- */
section[data-testid="stSidebar"] {
  background: linear-gradient(178deg, rgba(255,255,255,0.96) 0%, rgba(240,196,217,0.28) 100%);
  border-right: 2px solid rgba(152,160,172,0.32);
  backdrop-filter: blur(10px);
}
section[data-testid="stSidebar"] h3 { font-size: 1.5rem !important; }
section[data-testid="stSidebar"] label { font-weight: 600 !important; font-size: 0.9rem !important; }

/* ---------- ZAKLADKI ---------- */
.stTabs [data-baseweb="tab-list"] { gap: 5px; border-bottom: 2px solid rgba(152,160,172,0.34); }
.stTabs [data-baseweb="tab"] {
  background: rgba(255,255,255,0.60); border-radius: 11px 11px 0 0;
  padding: 10px 24px; font-weight: 650; font-size: 0.94rem;
  color: var(--tekst-mniej); border: 1.5px solid rgba(152,160,172,0.30); border-bottom: none;
}
.stTabs [aria-selected="true"] {
  background: linear-gradient(158deg, rgba(155,123,200,0.24), rgba(212,115,158,0.19)) !important;
  color: var(--fiolet-ciemny) !important; border-color: rgba(107,78,150,0.42) !important;
}

/* ---------- KARTY ---------- */
div[data-testid="stVerticalBlockBorderWrapper"] {
  background: rgba(255,255,255,0.80);
  border: 1.5px solid rgba(152,160,172,0.36) !important;
  border-radius: 15px !important;
  box-shadow: 0 3px 20px rgba(107,78,150,0.09);
  backdrop-filter: blur(8px);
}

/* ---------- METRYKI ---------- */
div[data-testid="stMetric"] {
  background: linear-gradient(148deg, rgba(255,255,255,0.94), rgba(240,196,217,0.40));
  border: 1.5px solid rgba(184,150,63,0.32);
  border-radius: 13px; padding: 15px 18px;
  box-shadow: 0 2px 14px rgba(107,78,150,0.09);
}
div[data-testid="stMetricValue"] { color: var(--fiolet-ciemny) !important; font-weight: 700 !important; }
div[data-testid="stMetricLabel"] { color: var(--tekst-mniej) !important; font-size: 0.76rem !important;
  text-transform: uppercase; letter-spacing: 0.07em; font-weight: 700 !important; }

/* ---------- PRZYCISKI ---------- */
.stButton > button, .stDownloadButton > button {
  background: linear-gradient(102deg, var(--fiolet-ciemny) 0%, var(--fiolet) 48%, var(--roz) 100%);
  color: #fff !important; border: none; border-radius: 10px;
  padding: 0.56rem 1.6rem; font-weight: 650; letter-spacing: 0.02em;
  box-shadow: 0 3px 14px rgba(107,78,150,0.30);
  transition: transform 0.16s ease, box-shadow 0.16s ease;
}
.stButton > button:hover, .stDownloadButton > button:hover {
  transform: translateY(-1.5px); box-shadow: 0 6px 20px rgba(107,78,150,0.38);
}

/* ---------- POLA ---------- */
.stTextArea textarea, .stTextInput input {
  border-radius: 10px !important; border: 1.5px solid rgba(152,160,172,0.52) !important;
  background: rgba(255,255,255,0.90) !important;
  font-family: 'IBM Plex Mono', 'Consolas', monospace !important;
  font-size: 0.86rem !important; font-weight: 500 !important;
}
.stTextArea textarea:focus, .stTextInput input:focus {
  border-color: var(--fiolet-jasny) !important;
  box-shadow: 0 0 0 3px rgba(155,123,200,0.19) !important;
}

/* ---------- SEKWENCJE ---------- */
.stCode, pre, code {
  background: linear-gradient(101deg, rgba(240,196,217,0.30), rgba(228,232,238,0.62)) !important;
  border: 1.5px solid rgba(184,150,63,0.30) !important;
  border-radius: 8px !important; color: var(--tekst) !important;
  font-family: 'IBM Plex Mono', 'Consolas', monospace !important;
  font-weight: 600 !important; letter-spacing: 0.055em; font-size: 0.86rem !important;
}

div[data-baseweb="slider"] div[role="slider"] {
  background: linear-gradient(138deg, var(--fiolet), var(--roz)) !important;
  border: 2.5px solid #fff !important; box-shadow: 0 2px 8px rgba(107,78,150,0.36) !important;
}

div[data-testid="stDataFrame"] {
  border-radius: 12px; overflow: hidden;
  border: 1.5px solid rgba(152,160,172,0.36);
  box-shadow: 0 2px 16px rgba(107,78,150,0.08);
}

div[data-testid="stAlert"] { border-radius: 11px; border-left-width: 4px; font-weight: 500; }

.streamlit-expanderHeader, details summary {
  background: rgba(255,255,255,0.70) !important; border-radius: 9px !important;
  border: 1.5px solid rgba(152,160,172,0.32) !important; font-weight: 650 !important;
}

hr { border-color: rgba(184,150,63,0.34) !important; }

/* ---------- ELEMENTY WLASNE ---------- */
.podtytul {
  font-family: 'IBM Plex Sans', sans-serif; font-size: 0.82rem;
  font-weight: 600; color: var(--tekst-mniej); margin: 0.1rem 0 0.5rem 0;
  text-transform: uppercase; letter-spacing: 0.14em;
}

/* pasek zasad pod naglowkiem */
.pasek-zasad {
  display: flex; gap: 3px; margin: 0.5rem 0 1.7rem 0; align-items: center;
}
.zasada {
  width: 25px; height: 25px; border-radius: 6px;
  display: flex; align-items: center; justify-content: center;
  font-family: 'IBM Plex Mono', 'Consolas', monospace; font-weight: 700; font-size: 0.78rem;
  color: #fff; box-shadow: 0 1px 5px rgba(107,78,150,0.22);
}
.z-A { background: linear-gradient(140deg, #6B4E96, #9B7BC8); }
.z-U { background: linear-gradient(140deg, #D4739E, #E8A5C2); }
.z-G { background: linear-gradient(140deg, #B8963F, #D9BC72); }
.z-C { background: linear-gradient(140deg, #7A8494, #A8B0BC); }
.linia-zasad { flex: 1; height: 2.5px; margin-left: 8px;
  background: linear-gradient(90deg, rgba(107,78,150,0.42), rgba(212,115,158,0.34), rgba(184,150,63,0.24), transparent); }

/* plakietki */
.plakietka {
  display: inline-block; padding: 4px 13px; border-radius: 10px;
  font-size: 0.71rem; font-weight: 700; letter-spacing: 0.06em;
  text-transform: uppercase; margin-right: 6px; margin-bottom: 4px;
}
.p-zloto  { background: rgba(184,150,63,0.20);  color: #7A6320; border: 1.5px solid rgba(184,150,63,0.42); }
.p-fiolet { background: rgba(107,78,150,0.17);  color: var(--fiolet-ciemny); border: 1.5px solid rgba(107,78,150,0.36); }
.p-roz    { background: rgba(212,115,158,0.19); color: #9C4670; border: 1.5px solid rgba(212,115,158,0.42); }
.p-srebro { background: rgba(152,160,172,0.24); color: #4E5766; border: 1.5px solid rgba(152,160,172,0.46); }

/* naglowek sekcji z ikona */
.sekcja { display: flex; align-items: center; gap: 11px; margin: 0.4rem 0 0.7rem 0; }
.sekcja svg { flex-shrink: 0; }
.sekcja-tytul { font-family: 'Spectral', 'Palatino Linotype', Georgia, serif; font-weight: 700;
  font-size: 1.28rem; color: var(--fiolet); letter-spacing: -0.008em; }

.stopka {
  text-align: center; color: var(--tekst-mniej); font-size: 0.82rem;
  padding: 2.4rem 0 1rem 0; font-family: 'Spectral', 'Palatino Linotype', Georgia, serif;
  font-weight: 500; letter-spacing: 0.05em;
  border-top: 1.5px solid rgba(184,150,63,0.26); margin-top: 2.4rem;
}
</style>
"""

# ---------- IKONY SVG ----------
IKONA = {
 'helisa': """<svg width="26" height="26" viewBox="0 0 26 26" fill="none">
   <path d="M6 1 C15 7, 15 12, 6 18 C-3 24, -3 25, 6 25" stroke="#6B4E96" stroke-width="2"/>
   <path d="M20 1 C11 7, 11 12, 20 18 C29 24, 29 25, 20 25" stroke="#9B7BC8" stroke-width="2"/>
   <path d="M8 5 L18 5 M9 10 L17 10 M9 15 L17 15 M8 20 L18 20" stroke="#D4739E" stroke-width="1.6"/>
 </svg>""",
 'tarcza': """<svg width="26" height="26" viewBox="0 0 26 26" fill="none">
   <path d="M13 1.5 L23 5.5 V13 C23 19, 18 23.5, 13 24.5 C8 23.5, 3 19, 3 13 V5.5 Z"
         stroke="#6B4E96" stroke-width="2" fill="rgba(155,123,200,0.13)"/>
   <path d="M8.5 13 L11.5 16.5 L17.5 9.5" stroke="#B8963F" stroke-width="2.4" stroke-linecap="round"/>
 </svg>""",
 'siatka': """<svg width="26" height="26" viewBox="0 0 26 26" fill="none">
   <circle cx="13" cy="4.5" r="3" stroke="#6B4E96" stroke-width="2"/>
   <circle cx="4.5" cy="19" r="3" stroke="#D4739E" stroke-width="2"/>
   <circle cx="21.5" cy="19" r="3" stroke="#B8963F" stroke-width="2"/>
   <path d="M11 7 L6.5 16 M15 7 L19.5 16 M7.5 19 L18.5 19" stroke="#98A0AC" stroke-width="1.8"/>
 </svg>""",
 'plazmid': """<svg width="26" height="26" viewBox="0 0 26 26" fill="none">
   <ellipse cx="13" cy="13" rx="10.5" ry="10.5" stroke="#6B4E96" stroke-width="2"/>
   <path d="M13 2.5 A10.5 10.5 0 0 1 22.6 9.2" stroke="#B8963F" stroke-width="3.4" stroke-linecap="round"/>
   <path d="M4.6 18.6 A10.5 10.5 0 0 1 3.1 10.4" stroke="#D4739E" stroke-width="3.4" stroke-linecap="round"/>
   <circle cx="13" cy="13" r="2.4" fill="rgba(107,78,150,0.28)"/>
 </svg>""",
 'ksiazka': """<svg width="26" height="26" viewBox="0 0 26 26" fill="none">
   <path d="M3.5 4 H10 C11.5 4, 13 5.2, 13 6.8 V22 C13 20.7, 11.5 19.8, 10 19.8 H3.5 Z"
         stroke="#6B4E96" stroke-width="2" fill="rgba(155,123,200,0.10)"/>
   <path d="M22.5 4 H16 C14.5 4, 13 5.2, 13 6.8 V22 C13 20.7, 14.5 19.8, 16 19.8 H22.5 Z"
         stroke="#D4739E" stroke-width="2" fill="rgba(212,115,158,0.10)"/>
 </svg>""",
}


def naglowek_sekcji(ikona: str, tytul: str) -> None:
    st.markdown(
        f'<div class="sekcja">{IKONA[ikona]}'
        f'<span class="sekcja-tytul">{tytul}</span></div>',
        unsafe_allow_html=True)


st.markdown(STYL, unsafe_allow_html=True)


# ============================================================================
# BUFOROWANIE
# ============================================================================

@st.cache_data(show_spinner=False)
def skanuj_offtarget(sciezka: str, guides_tuple: tuple, mrna_celu: str):
    """
    Skanowanie off-target. Buforowane po sciezce pliku i zestawie zapytan,
    wiec ponowne przeliczenie nastepuje tylko przy realnej zmianie danych.
    """
    guides = dict(guides_tuple)
    skaner = offtarget.OffTargetScanner(guides, mrna_celu=mrna_celu)
    pasek = st.progress(0.0, text='Scanning transcriptome...')
    szac = offtarget.szacuj_czas(sciezka)
    oczekiwane = max(szac['szac_nt_mln'] * 1000, 1)

    def postep(n_tx, nazwa):
        frakcja = min(n_tx / max(oczekiwane / 1500, 1), 0.99)
        pasek.progress(frakcja,
                       text=f'Scanning transcriptome - {n_tx:,} transcripts')

    wynik = skaner.scan(sciezka, progress=postep)
    pasek.empty()
    return wynik


@st.cache_data(show_spinner='Designing candidates...')
def generuj(mrna, dlugosci, gc_min, gc_max, min_asym, max_mfe, w5, w3):
    return design.generuj_kandydatow(
        mrna, dlugosci=dlugosci, gc_min=gc_min, gc_max=gc_max,
        min_asymetria=min_asym, max_mfe_guide=max_mfe,
        wyklucz_5prim_nt=w5, wyklucz_3prim_nt=w3, verbose=False)


@st.cache_data(show_spinner='Analysing conservation...')
def analizuj_konserwatywnosc(sciezka, dlugosci, prog):
    return conservation.analiza_konserwatywnosci(
        sciezka, dlugosci=dlugosci, prog_pokrycia=prog, verbose=False)


def parsuj_fasta(tekst):
    naglowek, buf = 'sekwencja', []
    for l in tekst.strip().splitlines():
        l = l.strip()
        if l.startswith('>'):
            naglowek = l[1:].split()[0] if len(l) > 1 else 'sekwencja'
        else:
            buf.append(l)
    return naglowek, ''.join(buf).upper().replace('U', 'T').replace(' ', '')


def zbuduj_z_rejestru(nazwa, guide_rna, promotor, klonowanie, petla_klucz):
    """
    Buduje kasete, przyjmujac OBIEKTY promotora i metody klonowania zamiast
    kluczy z rejestru. Pozwala uzyc elementow zdefiniowanych przez
    uzytkownika w trakcie sesji.

    Tymczasowo rejestruje obiekty w slownikach modulu constructs, wywoluje
    zbuduj_kasete, po czym przywraca stan poczatkowy.
    """
    kl_p = f'__tmp_{promotor.nazwa}'
    kl_k = f'__tmp_{klonowanie.nazwa}'
    constructs.PROMOTORY[kl_p] = promotor
    constructs.KLONOWANIE[kl_k] = klonowanie
    try:
        return constructs.zbuduj_kasete(
            nazwa=nazwa, guide_rna=guide_rna,
            promotor_klucz=kl_p, metoda_klonowania_klucz=kl_k,
            petla_klucz=petla_klucz)
    finally:
        constructs.PROMOTORY.pop(kl_p, None)
        constructs.KLONOWANIE.pop(kl_k, None)


def do_tsv(dane, kolumny, sep='\t') -> str:
    buf = io.StringIO()
    buf.write(sep.join(kolumny) + '\n')
    for c in dane:
        buf.write(sep.join(str(c.get(k, '')) for k in kolumny) + '\n')
    return buf.getvalue()


def do_xlsx(grupy, kolumny) -> bytes:
    """
    Buduje skoroszyt XLSX. `grupy` to slownik {nazwa_arkusza: lista_wierszy}.
    Kazda grupa trafia na osobny arkusz.

    Wymaga openpyxl. Zwraca None, jesli biblioteka niedostepna.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    naglowek_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    naglowek_fill = PatternFill('solid', fgColor='6B4E96')
    tresc_font = Font(name='Arial', size=10)
    mono_font = Font(name='Consolas', size=10)
    kol_sekwencji = {'guide_rna', 'passenger_rna', 'seed_2_8'}

    for nazwa_ark, wiersze in grupy.items():
        ws = wb.create_sheet(str(nazwa_ark)[:31])
        for j, k in enumerate(kolumny, 1):
            c = ws.cell(row=1, column=j, value=k)
            c.font = naglowek_font
            c.fill = naglowek_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        for i, w in enumerate(wiersze, 2):
            for j, k in enumerate(kolumny, 1):
                v = w.get(k, '')
                if isinstance(v, bool):
                    v = 'tak' if v else 'nie'
                c = ws.cell(row=i, column=j, value=v)
                c.font = mono_font if k in kol_sekwencji else tresc_font
        ws.freeze_panes = 'A2'
        for j, k in enumerate(kolumny, 1):
            szer = max([len(str(k))] +
                       [len(str(w.get(k, ''))) for w in wiersze[:200]] or [8])
            ws.column_dimensions[get_column_letter(j)].width = min(szer + 2, 40)

    bufor = io.BytesIO()
    wb.save(bufor)
    return bufor.getvalue()


def zapisz_tymczasowo(plik) -> str:
    with tempfile.NamedTemporaryFile(mode='w', suffix='.fasta',
                                     delete=False) as tf:
        tf.write(plik.getvalue().decode('utf-8'))
        return tf.name


# ============================================================================
# PASEK BOCZNY
# ============================================================================

st.sidebar.markdown('### Settings')

profil_nazwa = st.sidebar.selectbox(
    'Target organism', sorted(hosts.PROFILE),
    index=sorted(hosts.PROFILE).index('plant'))
profil = hosts.get_profile(profil_nazwa)

st.sidebar.markdown(
    f'<span class="plakietka p-fiolet">GC {profil.gc_min:.0f}–{profil.gc_max:.0f}%</span>'
    f'<span class="plakietka p-roz">5′ {profil.guide_5_dozwolone or "dowolny"}</span>'
    f'<span class="plakietka p-zloto">{len(profil.motywy_zabronione)} motifs</span>',
    unsafe_allow_html=True)
st.sidebar.caption(profil.opis)

dlugosci = st.sidebar.multiselect(
    'siRNA lengths (nt)', [19, 20, 21, 22, 23, 24],
    default=list(profil.dlugosci))

st.sidebar.divider()

zaawansowany = st.sidebar.toggle(
    'Advanced mode', value=False,
    help='Unlocks manual threshold editing. By default the values come '
         'from the organism profile and need no adjustment.')

if zaawansowany:
    st.sidebar.warning('Changing thresholds without biological '
                       'justification reduces comparability of results.')
    gc_min, gc_max = st.sidebar.slider('GC content (%)', 20.0, 80.0,
                                       (profil.gc_min, profil.gc_max), 1.0)
    min_asym = st.sidebar.slider('Min. asymmetry (kcal/mol)', 0.0, 3.0,
                                 profil.min_asymetria, 0.1)
    max_mfe = st.sidebar.slider('Max. guide-strand MFE', -10.0, 0.0,
                                profil.max_mfe_guide, 0.5)
    c1, c2 = st.sidebar.columns(2)
    w5 = c1.number_input('Exclude 5-prime', 0, 300, profil.wyklucz_5prim_nt, 5)
    w3 = c2.number_input('Exclude 3-prime', 0, 300, profil.wyklucz_3prim_nt, 5)
    with st.sidebar.expander('Scoring weights'):
        wagi = {n: st.slider(s['opis'][:42], 0.0, 1.0, s['waga'], 0.05,
                             key=f'w_{n}')
                for n, s in scoring.METRYKI.items()}
        suma = sum(wagi.values())
        wagi = {k: v / suma for k, v in wagi.items()} if suma > 0 else None
else:
    gc_min, gc_max = profil.gc_min, profil.gc_max
    min_asym, max_mfe = profil.min_asymetria, profil.max_mfe_guide
    w5, w3 = profil.wyklucz_5prim_nt, profil.wyklucz_3prim_nt
    wagi = None
    st.sidebar.caption('Thresholds set automatically from the organism '
                       'profile.')

st.sidebar.divider()
if not design.VIENNA_DOSTEPNE:
    st.sidebar.error('ViennaRNA unavailable - reduced mode.')
else:
    st.sidebar.success('ViennaRNA active')


# ============================================================================
# NAGLOWEK
# ============================================================================

st.markdown('# siRNA Design Studio')
st.markdown('<div class="podtytul">Design, evaluation and construction of '
            'interfering RNA molecules</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="pasek-zasad">'
    + ''.join(f'<div class="zasada z-{b}">{b}</div>'
              for b in 'AUGCAUGCUAGC')
    + '<div class="linia-zasad"></div></div>',
    unsafe_allow_html=True)

zakl = st.tabs(['Sequence', 'Data sources', 'Conservation', 'Candidates',
                'Constructs', 'Order sheet', 'Methods'])


# --- 1. SEKWENCJA ------------------------------------------------------------

with zakl[0]:
    kol1, kol2 = st.columns([3, 2], gap='large')

    with kol1:
        naglowek_sekcji('helisa', 'Target sequence')
        zrodlo = st.radio('Source', ['Paste', 'FASTA file'],
                          horizontal=True, label_visibility='collapsed')
        tekst = ''
        if zrodlo == 'Paste':
            tekst = st.text_area('mRNA / ORF', height=210,
                                 placeholder='>my_gene\nATGGTGAGCAAGGGC...',
                                 label_visibility='collapsed')
        else:
            p = st.file_uploader('FASTA', ['fasta', 'fa', 'txt'],
                                 label_visibility='collapsed')
            if p:
                tekst = p.read().decode('utf-8')

        if tekst:
            nag, mrna = parsuj_fasta(tekst)
            zle = set(mrna) - set('ACGTN')
            if zle:
                st.error(f'Invalid characters: {sorted(zle)}')
            else:
                st.success(f'**{nag}** — {len(mrna)} nt')
                st.session_state['mrna'] = mrna
                st.session_state['naglowek'] = nag

    with kol2:
        naglowek_sekcji('siatka', 'Reference data')

        with st.container(border=True):
            st.markdown('#### Host transcriptome')
            st.caption('For off-target analysis. cDNA file of the species in '
                       'which the siRNA is to act.')
            plik_tx = st.file_uploader('cDNA FASTA', ['fasta', 'fa'],
                                       key='upload_tx', label_visibility='collapsed')
            if plik_tx:
                st.session_state['sciezka_tx'] = zapisz_tymczasowo(plik_tx)
                st.success(plik_tx.name)
                if OFFTARGET_OK:
                    sz = offtarget.szacuj_czas(
                        st.session_state['sciezka_tx'])
                    st.info(
                        f'Size {sz["rozmiar_mb"]} MB - scanning will take '
                        f'about **{sz["szac_minut"]} min**. The result is '
                        f'cached and computed once per session.')
                    if sz['rozmiar_mb'] > 90:
                        st.warning(
                            'The file size suggests this may be a **genome** '
                            'rather than a transcriptome. Genomes contain '
                            'introns and intergenic regions that are never '
                            'transcribed, so an siRNA will never encounter '
                            'them. Hits in those regions are biologically '
                            'irrelevant and lengthen the analysis. A **cDNA** '
                            'file is recommended - about 50 MB for '
                            '*Arabidopsis*, against 119 MB for the genome.')
            st.caption('Sources and automated download: see the '
                       '**Data sources** tab.')


        with st.container(border=True):
            st.markdown('#### Isolate set')
            st.caption('For conservation analysis. Multiple sequences of the '
                       'same gene or genome from different isolates.')
            plik_izo = st.file_uploader('Multi-FASTA', ['fasta', 'fa'],
                                        key='upload_izo',
                                        label_visibility='collapsed')
            if plik_izo:
                st.session_state['sciezka_izolaty'] = zapisz_tymczasowo(plik_izo)
                st.success(plik_izo.name)


# --- 2. KONSERWATYWNOSC ------------------------------------------------------

with zakl[1]:
    naglowek_sekcji('siatka', 'Data sources')
    st.caption('Where reference data comes from, and how to obtain it.')

    d1, d2 = st.tabs(['Manual download', 'Automated retrieval (NCBI)'])

    with d1:
        st.markdown("""
### Host transcriptomes

Required for off-target analysis. Download the **cDNA** file for your species,
unpack it if compressed, and upload it in the **Sequence** tab.

| Species | Source | File to look for |
|---|---|---|
| *Arabidopsis thaliana* | [TAIR](https://www.arabidopsis.org/download/) | `TAIR10_cdna_*` |
| Tomato, potato, pepper | [Sol Genomics Network](https://solgenomics.net/ftp/) | `ITAG*_cDNA.fasta` |
| Other plants | [Ensembl Plants](https://plants.ensembl.org/info/data/ftp/index.html) | cDNA column, `*.cdna.all.fa.gz` |
| Human, mouse, others | [Ensembl](https://www.ensembl.org/info/data/ftp/index.html) | cDNA column, `*.cdna.all.fa.gz` |

**Use cDNA, not the genome.** A genome contains introns and intergenic
regions that are never transcribed, so an siRNA will never encounter them.
Hits in those regions are biologically irrelevant and roughly double the
analysis time. For *Arabidopsis* the cDNA file is about 50 MB against 119 MB
for the genome.

Files ending in `.gz` must be unpacked before upload — 7-Zip on Windows,
a double click on macOS.

### Isolate sets

Required for conservation analysis. A multi-FASTA containing many sequences
of the same gene or genome from different isolates.

- [NCBI Nucleotide](https://www.ncbi.nlm.nih.gov/nuccore/) — search the
  organism name, then **Send to → File → FASTA**
- For a bounded, query-defined set the automated route in the next tab is
  faster and records what was filtered out
""")

    with d2:
        if not NCBI_OK:
            st.error('Module ncbi.py unavailable.')
        else:
            st.markdown("""
Retrieval through the NCBI **E-utilities** interface. Appropriate for bounded,
query-defined sets: viral isolates, a specific gene, a handful of reference
transcripts.

**Not appropriate for whole transcriptomes.** Retrieving tens of thousands of
records one request at a time is slow, loads a public service unreasonably and
duplicates files published as single downloads. Use the previous tab for those.
""")
            st.divider()

            n1, n2 = st.columns([2, 1])
            email = n1.text_input(
                'E-mail address (required)', '',
                placeholder='you@institution.org',
                help='NCBI uses this to contact clients generating excessive '
                     'load, and blocks unidentified traffic. It is a '
                     'condition of use, not a courtesy.')
            klucz = n2.text_input(
                'API key (optional)', '', type='password',
                help='Raises the limit from 3 to 10 requests per second. '
                     'Free, from an NCBI account.')

            szablon = st.selectbox(
                'Query template', sorted(ncbi.QUERY_TEMPLATES),
                format_func=lambda k: (f'{k} — '
                                       f'{ncbi.QUERY_TEMPLATES[k]["description"][:60]}'))
            spec = ncbi.QUERY_TEMPLATES[szablon]
            st.caption(spec['description'])
            if 'example' in spec:
                st.caption(f'Example input: {spec["example"]}')

            pola = re.findall(r'\{(\w+)\}', spec['template'])
            wartosci = {}
            if pola:
                kols = st.columns(len(pola))
                for kol, p in zip(kols, pola):
                    wartosci[p] = kol.text_input(p.replace('_', ' '), '')

            zapytanie = ''
            if all(wartosci.get(p) for p in pola):
                zapytanie = spec['template'].format(**wartosci)
                st.code(zapytanie, language=None)

            st.markdown('**Quality filters**')
            q1, q2, q3, q4 = st.columns(4)
            maks = q1.number_input('Max records', 10, 5000, 310, 10)
            min_dl = q2.number_input('Min length (nt)', 0, 100000, 0, 100)
            maks_dl = q3.number_input('Max length (nt)', 0, 1000000, 0, 100)
            wyklucz = q4.checkbox('Exclude partial', True)
            st.caption('Partial records lower apparent conservation at their '
                       'ends because the region is absent, not because it is '
                       'variable.')

            c_a, c_b = st.columns(2)
            if c_a.button('Count matching records', use_container_width=True,
                          disabled=not (email and zapytanie)):
                try:
                    cli = ncbi.NCBIClient(email=email, api_key=klucz or None)
                    st.info(f'{cli.count(spec["db"], zapytanie)} records match '
                            f'this query.')
                except Exception as e:
                    st.error(f'{e}')

            if c_b.button('Retrieve sequences', type='primary',
                          use_container_width=True,
                          disabled=not (email and zapytanie)):
                pasek = st.progress(0.0, text='Contacting NCBI...')
                try:
                    cli = ncbi.NCBIClient(email=email, api_key=klucz or None)
                    stan = {'n': 0}

                    def post(msg):
                        stan['n'] += 1
                        pasek.progress(min(stan['n'] / 30, 0.95), text=msg)

                    recs, rap = cli.fetch_sequences(
                        spec['db'], zapytanie, max_records=int(maks),
                        min_length=int(min_dl) or None,
                        max_length=int(maks_dl) or None,
                        exclude_partial=wyklucz, progress=post)
                    pasek.empty()

                    if recs:
                        sciezka = tempfile.NamedTemporaryFile(
                            mode='w', suffix='.fasta', delete=False).name
                        ncbi.NCBIClient.save_fasta(recs, sciezka)
                        st.session_state['sciezka_izolaty'] = sciezka
                        st.success(f'{len(recs)} sequences retrieved and set '
                                   f'as the isolate set.')
                        st.caption(rap.summary())
                        for uw in rap.notes:
                            st.caption(uw)
                        with open(sciezka) as fh:
                            st.download_button('Download FASTA', fh.read(),
                                               'ncbi_sequences.fasta')
                    else:
                        st.warning(rap.summary())
                except Exception as e:
                    pasek.empty()
                    st.error(f'Retrieval failed: {e}')


with zakl[2]:
    naglowek_sekcji('tarcza', 'Conservation analysis')
    st.caption('Identifies regions present in the largest fraction of '
               'isolates. An siRNA directed at such a region acts on the '
               'whole population and leaves the pathogen fewer routes of '
               'escape by mutation.')

    if not KONSERWATYWNOSC_OK:
        st.error('Module conservation.py unavailable.')
    elif 'sciezka_izolaty' not in st.session_state:
        st.info('Upload an isolate set in the **Sequence** tab to run '
                'this analysis.')
        with st.expander('How it works, and why without alignment'):
            st.markdown("""
The classical approach performs a multiple sequence alignment and then
computes Shannon entropy for each column. This requires an external program
(MAFFT) and is costly for several hundred genomes.

This module uses a direct method: for every window in the reference sequence
it determines **the fraction of the remaining isolates that contain exactly
the same string**.

The rationale is biological. An siRNA requires near-perfect complementarity
— a single mismatch in the seed region is enough for the target not to be
recognised. What matters is therefore not conservation in the evolutionary
sense, but the presence of one specific k-mer in a given isolate. That
quantity is measured directly, without alignment.

**Limitation.** The method detects substitutions only. Insertions and
deletions would require an alignment; Shannon entropy is implemented in the
module and available for input that has already been aligned.
""")
    else:
        auto = st.toggle(
            'Determine the threshold from the data', value=True,
            help='Locates the knee of the coverage curve rather than '
                 'assuming a conventional value. The point of maximum '
                 'curvature marks the transition between variable and '
                 'conserved windows.')

        if auto:
            if st.button('Optimise threshold'):
                st.session_state['opt'] = conservation.optimise_threshold(
                    st.session_state['sciezka_izolaty'],
                    dlugosc=sorted(dlugosci)[0], verbose=False)
            if 'opt' in st.session_state:
                o = st.session_state['opt']
                kolor = {'high': 'p-zloto', 'medium': 'p-roz',
                         'low': 'p-srebro'}[o['confidence']]
                st.markdown(
                    f'<span class="plakietka p-fiolet">threshold '
                    f'{o["threshold"]:.2f}</span>'
                    f'<span class="plakietka {kolor}">confidence '
                    f'{o["confidence"]}</span>'
                    f'<span class="plakietka p-srebro">{o["n_windows"]} '
                    f'windows</span>', unsafe_allow_html=True)
                st.caption(f'Method: {o["method"]}')
                with st.expander('Diagnostics and curve'):
                    for d in o['diagnostics']:
                        st.write(f'- {d}')
                        st.line_chart(
                            {'windows above threshold':
                             {str(t): n for t, n in o['curve']}})
                        break
                    for d in o['diagnostics'][1:]:
                        st.write(f'- {d}')
                prog = o['threshold']
            else:
                prog = 0.95
                st.caption('Press the button to compute the threshold.')
        else:
            c1, _ = st.columns([1, 2])
            prog = c1.slider('Coverage threshold', 0.50, 1.00, 0.95, 0.01,
                                help='Minimum fraction of isolates containing '
                                   'the given fragment.')

        if st.button('Run analysis', type='primary'):
            st.session_state['kons'] = analizuj_konserwatywnosc(
                st.session_state['sciezka_izolaty'], tuple(sorted(dlugosci)), prog)

        if 'kons' in st.session_state:
            w = st.session_state['kons']
            bloki = conservation.scal_regiony(w['regiony'], min_dlugosc=30)
            m1, m2, m3 = st.columns(3)
            m1.metric('Isolates', w['n_sekwencji'])
            m2.metric('Windows above threshold', len(w['regiony']))
            m3.metric('Contiguous blocks >=30 nt', len(bloki))

            st.caption(f'Reference: **{w["referencja"][0]}** '
                       f'({len(w["referencja"][1])} nt)')

            if bloki:
                st.markdown('#### Conserved blocks')
                st.caption('Longer stretches in which every window meets '
                           'the threshold, leaving freedom in the choice of '
                           'exact position.')
                st.dataframe(bloki[:20], use_container_width=True)

            if w['regiony']:
                st.markdown('#### Best windows')
                st.dataframe(w['regiony'][:40], use_container_width=True,
                             height=300)

            st.session_state['filtruj_kons'] = st.checkbox(
                'Restrict candidates to conserved regions', value=True)


# --- 3. KANDYDACI ------------------------------------------------------------

with zakl[3]:
    if 'mrna' not in st.session_state:
        st.info('Enter a sequence in the **Sequence** tab first.')
    elif not dlugosci:
        st.warning('Select at least one siRNA length.')
    else:
        mrna = st.session_state['mrna']

        # ---------- PANEL URUCHOMIENIA ----------
        with st.container(border=True):
            u1, u2 = st.columns([2, 1])
            with u1:
                st.markdown('#### Analysis scope')
                ma_tx = 'sciezka_tx' in st.session_state
                st.markdown(
                    f'<span class="plakietka p-fiolet">{len(mrna)} nt</span>'
                    f'<span class="plakietka p-roz">'
                    f'{", ".join(str(d) for d in sorted(dlugosci))} nt</span>'
                    f'<span class="plakietka '
                    f'{"p-zloto" if ma_tx else "p-srebro"}">'
                    f'off-target: {"yes" if ma_tx else "skipped"}</span>',
                    unsafe_allow_html=True)
                if ma_tx and OFFTARGET_OK:
                    sz = offtarget.szacuj_czas(st.session_state['sciezka_tx'])
                    st.caption(f'Scanning the transcriptome '
                               f'({sz["rozmiar_mb"]} MB) will take about '
                               f'{sz["szac_minut"]} min. The result is '
                               f'cached and computed once.')
                else:
                    st.caption('Without a transcriptome the off-target '
                               'analysis is skipped and the ranking does not '
                               'account for sequence safety.')
            with u2:
                st.write('')
                st.write('')
                uruchom = st.button('Run analysis', type='primary',
                                    use_container_width=True)

        if uruchom:
            st.session_state.pop('ranking', None)

            pasek = st.progress(0.0, text='Generating candidates...')
            kand = generuj(mrna, tuple(sorted(dlugosci)), gc_min, gc_max,
                           min_asym, max_mfe, w5, w3)
            pasek.progress(0.25, text=f'{len(kand)} candidates generated. '
                                      f'Applying filters...')

            odrzuc_motywy = odrzuc_kons = 0
            if profil.motywy_zabronione:
                przed = len(kand)
                kand = [k for k in kand
                        if not any(m in k['guide_rna']
                                   for m in profil.motywy_zabronione)]
                odrzuc_motywy = przed - len(kand)

            if (st.session_state.get('filtruj_kons')
                    and 'kons' in st.session_state):
                przed = len(kand)
                kand = conservation.filtruj_kandydatow(
                    kand, st.session_state['kons'], prog=0.95)
                odrzuc_kons = przed - len(kand)

            if not kand:
                pasek.empty()
                st.error('No candidates passed the filters. In advanced '
                         'mode, widen the GC range or lower the asymmetry '
                         'requirement.')
            else:
                pasek.progress(0.40, text='Off-target analysis...')
                if 'sciezka_tx' in st.session_state and OFFTARGET_OK:
                    gt = tuple(sorted((k['nazwa'], k['guide_rna'])
                                      for k in kand))
                    rap = skanuj_offtarget(st.session_state['sciezka_tx'],
                                           gt, mrna)
                    for k in kand:
                        r = rap[k['nazwa']]
                        k['koszt_offtarget'] = (r['min_koszt']
                                                if r['min_koszt'] is not None
                                                else 30.0)
                        k['seed_czystosc'] = r['n_8mer']
                        k['offtarget_flaga'] = r['flaga']
                        k['trafien_offtarget'] = r['n_trafien_pelnych']
                        k['trafien_krytycznych'] = r['n_krytycznych']
                        k['wzbogacenie_seed'] = r['wzbogacenie_8mer']
                        k['trafien_zamierzonych'] = r['n_trafien_zamierzonych']
                    ot_wykonane = True
                else:
                    for k in kand:
                        k['koszt_offtarget'] = 20.0
                        k['seed_czystosc'] = 0
                        k['offtarget_flaga'] = 'NOT CHECKED'
                        k['trafien_offtarget'] = 0
                        k['trafien_krytycznych'] = 0
                        k['wzbogacenie_seed'] = None
                        k['trafien_zamierzonych'] = 0
                    ot_wykonane = False

                pasek.progress(0.80, text='Scoring and ranking...')
                wyn = scoring.normalize_and_score(kand, wagi=wagi)
                wraz = scoring.sensitivity_analysis(kand, top_n=5)
                par = {c['nazwa'] for c in scoring.pareto_front(kand)}
                for c in wyn:
                    c['pareto'] = c['nazwa'] in par

                pasek.progress(1.0, text='Done')
                pasek.empty()

                st.session_state['ranking'] = wyn
                st.session_state['wrazliwosc'] = wraz
                st.session_state['ot_wykonane'] = ot_wykonane
                st.session_state['odrzucone'] = {
                    'motywy': odrzuc_motywy, 'konserwatywnosc': odrzuc_kons}

        # ---------- WYNIKI ----------
        if 'ranking' in st.session_state:
            wyn = st.session_state['ranking']
            wraz = st.session_state['wrazliwosc']
            ot_wykonane = st.session_state['ot_wykonane']
            odrz = st.session_state['odrzucone']

            n_ok = sum(1 for c in wyn if c['offtarget_flaga'] == 'OK')
            n_sprawdz = sum(1 for c in wyn
                            if c['offtarget_flaga'].startswith('REVIEW'))
            n_odrzuc = sum(1 for c in wyn
                           if c['offtarget_flaga'].startswith('REJECT'))

            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric('Candidates', len(wyn))
            m2.metric('Pareto front', len(scoring.pareto_front(wyn)))
            m3.metric('Off-target OK', n_ok,
                      delta=None if ot_wykonane else 'not checked',
                      delta_color='off')
            m4.metric('To review', n_sprawdz)
            m5.metric('Weight stability', f'{wraz["stabilnosc"]:.2f}')

            if not ot_wykonane:
                st.warning('Off-target analysis was not performed - no '
                           'transcriptome was uploaded.')
            elif n_odrzuc == len(wyn):
                st.error(
                    'Every candidate carries a critical hit - at most three '
                    'mismatches with an intact seed region. Under so strict '
                    'a criterion this is unusual. Check whether the target '
                    'gene is present in the database under a sequence that '
                    'differs from the one pasted (a different transcript '
                    'variant, for instance), which would prevent it from '
                    'being recognised as the intended target.')

            if odrz['motywy'] or odrz['konserwatywnosc']:
                st.caption(
                    f'Filtered out before scoring: '
                    f'{odrz["motywy"]} by forbidden motifs, '
                    f'{odrz["konserwatywnosc"]} by conservation.')
            st.caption(f'Weight sensitivity analysis: '
                       f'{wraz["interpretacja"]}')

            st.divider()

            # ---------- FILTRY TABELI ----------
            st.markdown('#### Result filtering')
            f1, f2, f3, f4 = st.columns(4)
            f_ot = f1.multiselect(
                'Off-target status',
                ['OK', 'REVIEW', 'REJECT', 'NOT CHECKED'],
                default=['OK', 'REVIEW', 'NOT CHECKED'],
                help='Only candidates with a critical hit are rejected: at most '
                     '3 mismatches with an intact seed region.')
            f_par = f2.selectbox('Pareto front',
                                 ['all', 'front only'])
            f_dl = f3.multiselect('Length', sorted(dlugosci),
                                  default=sorted(dlugosci))
            f_min = f4.slider('Minimum score', 0.0, 1.0, 0.0, 0.05)

            def pasuje(c):
                st_ot = c['offtarget_flaga']
                kat = ('OK' if st_ot == 'OK'
                       else 'NOT CHECKED' if st_ot == 'NOT CHECKED'
                       else 'REVIEW' if st_ot.startswith('REVIEW')
                       else 'REJECT')
                if kat not in f_ot:
                    return False
                if f_par == 'front only' and not c.get('pareto'):
                    return False
                if c['dlugosc'] not in f_dl:
                    return False
                if c['wynik'] < f_min:
                    return False
                return True

            widoczne = [c for c in wyn if pasuje(c)]
            st.caption(f'Showing **{len(widoczne)}** of {len(wyn)}')

            widok = st.radio(
                'Presentation',
                ['Combined ranking', 'Grouped by length', 'Nested series'],
                horizontal=True,
                help='Grouping gives each length class its own ranking. This '
                     'matters when comparing 21, 22 and 24 nt variants: each '
                     'enters a different DCL pathway, so they do not compete '
                     'directly.')

            if not widoczne:
                st.info('No candidate meets the selected criteria.')
            else:
                kol = ['ranga', 'nazwa', 'wynik', 'pareto', 'dlugosc',
                       'poz_start_1based', 'guide_rna', 'seed_2_8',
                       'gc_proc', 'asymetria', 'Tm', 'offtarget_flaga',
                       'trafien_krytycznych', 'wzbogacenie_seed',
                       'trafien_zamierzonych']
                if 'pokrycie_izolatow' in wyn[0]:
                    kol.insert(4, 'pokrycie_izolatow')

                if widok == 'Nested series':
                    st.caption(
                        'A nested series consists of candidates of different '
                        'lengths targeting the **same site**. This is the '
                        'only arrangement in which length is the sole '
                        'variable — if each length targets a different site, '
                        'the effect of length cannot be separated from the '
                        'effect of sequence context.')

                    tol = st.slider(
                        'Position tolerance (nt)', 0, 6, 3,
                        help='Windows of different lengths rarely begin at '
                             'exactly the same coordinate, because a window '
                             'starting one nucleotide earlier may fail a '
                             'filter. A tolerance of 2-3 nt is normal.')

                    serie = design.serie_zagniezdzone(widoczne,
                                                      tolerancja_poz=tol)
                    if not serie:
                        st.warning(
                            f'No nested series found within {tol} nt. Increase '
                            f'the tolerance, or select more length classes in '
                            f'the sidebar.')
                        grupy = {'all': widoczne}
                        st.dataframe(
                            [{k: c.get(k) for k in kol} for c in widoczne],
                            use_container_width=True, height=430)
                    else:
                        # --- podsumowanie wszystkich serii, posortowane ---
                        wszystkie_dl = sorted(dlugosci)
                        podsumowanie = []
                        for i, seria in enumerate(serie, 1):
                            dl_list = sorted(d for d in seria
                                             if isinstance(d, int))
                            sr = sum(seria[d]['wynik']
                                     for d in dl_list) / len(dl_list)
                            najgorszy = min(seria[d]['wynik']
                                            for d in dl_list)
                            statusy = {seria[d]['offtarget_flaga']
                                       for d in dl_list}
                            wiersz = {
                                'series': i,
                                'anchor_end': seria['_poz_kotwicy'],
                                'lengths': ', '.join(f'{d}' for d in dl_list),
                                'complete': len(dl_list) == len(wszystkie_dl),
                                'same_seed': seria['_seed_zgodny'],
                                'seed': seria.get('_seed') or 'differs',
                                'mean_score': round(sr, 4),
                                'lowest_score': round(najgorszy, 4),
                                'all_offtarget_OK': statusy == {'OK'},
                            }
                            podsumowanie.append((wiersz, seria))

                        # kolejnosc: kompletne -> zgodny seed -> off-target OK
                        # -> najwyzsza srednia
                        podsumowanie.sort(
                            key=lambda t: (not t[0]['complete'],
                                           not t[0]['same_seed'],
                                           not t[0]['all_offtarget_OK'],
                                           -t[0]['mean_score']))

                        st.success(f'{len(serie)} nested series found.')

                        st.markdown('#### Series ranked')
                        st.caption(
                            'Sorted by: completeness first (all length '
                            'classes present), then seed consistency, then '
                            'off-target status, then mean score. '
                            '**Mean score** is the arithmetic mean of the '
                            'composite scores of the variants in the series; '
                            '**lowest score** is the weakest member, which '
                            'matters because a series is only as usable as '
                            'its worst variant.')
                        st.dataframe([w for w, _ in podsumowanie],
                                     use_container_width=True,
                                     height=min(60 + 35 * len(podsumowanie),
                                                300))

                        gotowe = [w for w, _ in podsumowanie
                                  if w['complete'] and w['same_seed']
                                  and w['all_offtarget_OK']]
                        if gotowe:
                            n = gotowe[0]
                            st.info(
                                f'**Recommended: series {n["series"]}** '
                                f'(anchor end {n["anchor_end"]}, seed '
                                f'{n["seed"]}, mean score '
                                f'{n["mean_score"]}). It is the '
                                f'highest-scoring series that is complete, '
                                f'has a consistent seed across all lengths, '
                                f'and carries no off-target flag.')
                        else:
                            braki = []
                            if not any(w['complete'] for w, _ in podsumowanie):
                                braki.append('none is complete')
                            if not any(w['same_seed'] for w, _ in podsumowanie):
                                braki.append('none has a consistent seed')
                            if not any(w['all_offtarget_OK']
                                       for w, _ in podsumowanie):
                                braki.append('all carry an off-target flag')
                            st.warning(
                                'No series meets all three criteria — '
                                + ', '.join(braki) + '. Consider raising the '
                                'position tolerance, or relaxing the '
                                'off-target filter to include REVIEW.')

                        st.divider()
                        grupy = {}
                        for wiersz, seria in podsumowanie:
                            i = wiersz['series']
                            dl_list = sorted(d for d in seria
                                             if isinstance(d, int))
                            with st.container(border=True):
                                znaczniki = []
                                if wiersz['complete']:
                                    znaczniki.append(
                                        '<span class="plakietka p-zloto">'
                                        'complete</span>')
                                if wiersz['same_seed']:
                                    znaczniki.append(
                                        '<span class="plakietka p-fiolet">'
                                        f'seed {wiersz["seed"]}</span>')
                                else:
                                    znaczniki.append(
                                        '<span class="plakietka p-roz">'
                                        'seed differs</span>')
                                if wiersz['all_offtarget_OK']:
                                    znaczniki.append(
                                        '<span class="plakietka p-srebro">'
                                        'off-target OK</span>')
                                st.markdown(
                                    f'#### Series {i} — anchor end '
                                    f'{wiersz["anchor_end"]} · mean '
                                    f'{wiersz["mean_score"]:.3f} · lowest '
                                    f'{wiersz["lowest_score"]:.3f}')
                                st.markdown(''.join(znaczniki),
                                            unsafe_allow_html=True)
                                wiersze_s = [seria[d] for d in dl_list]
                                grupy[f'series{i}_end{wiersz["anchor_end"]}'] = \
                                    wiersze_s
                                st.dataframe(
                                    [{k: c.get(k) for k in kol}
                                     for c in wiersze_s],
                                    use_container_width=True,
                                    height=60 + 35 * len(wiersze_s))
                                brakuje = [d for d in sorted(dlugosci)
                                           if d not in seria]
                                if brakuje:
                                    st.caption(
                                        f'Missing length classes: '
                                        f'{", ".join(str(b) for b in brakuje)} '
                                        f'nt — no window of that length passed '
                                        f'the filters at this site.')
                elif widok == 'Combined ranking':
                    grupy = {'all': widoczne}
                    st.dataframe(
                        [{k: c.get(k) for k in kol} for c in widoczne],
                        use_container_width=True, height=430)
                else:
                    grupy = {}
                    for dl in sorted({c['dlugosc'] for c in widoczne}):
                        podzbior = [c for c in widoczne if c['dlugosc'] == dl]
                        # ranking wewnatrz grupy liczony od nowa
                        for i, c in enumerate(podzbior, 1):
                            c['ranga_w_grupie'] = i
                        grupy[f'{dl}nt'] = podzbior

                    opis_dcl = profil.opis_dlugosci
                    for nazwa_gr, podzbior in grupy.items():
                        dl = podzbior[0]['dlugosc']
                        st.markdown(f'#### {dl} nt - {len(podzbior)} candidates')
                        if dl in opis_dcl:
                            st.caption(opis_dcl[dl])
                        kol_gr = ['ranga_w_grupie'] + [k for k in kol
                                                       if k != 'ranga']
                        st.dataframe(
                            [{k: c.get(k) for k in kol_gr} for c in podzbior],
                            use_container_width=True,
                            height=min(60 + 35 * len(podzbior), 330))

                st.divider()
                st.markdown('#### Export')

                kol_eksport = (['ranga_w_grupie'] + [k for k in kol
                                                     if k != 'ranga']
                               if widok == 'Grouped by length' else kol)

                e1, e2, e3 = st.columns(3)

                e1.download_button(
                    'CSV - shown', do_tsv(widoczne, kol_eksport, sep=';'),
                    'siRNA_widoczne.csv', 'text/csv',
                    use_container_width=True,
                    help='Semicolon-separated, opens directly in Excel locales that '
                         'use the comma as decimal mark.')

                e2.download_button(
                    'TSV - shown', do_tsv(widoczne, kol_eksport),
                    'siRNA_widoczne.tsv', 'text/tab-separated-values',
                    use_container_width=True,
                    help='Format tabulatorowy, wygodny do dalszej analizy '
                         'w R lub Pythonie.')

                dane_xlsx = do_xlsx(grupy, kol_eksport)
                if dane_xlsx is not None:
                    e3.download_button(
                        'XLSX - sheets', dane_xlsx, 'siRNA_wyniki.xlsx',
                        'application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sheet',
                        use_container_width=True,
                        help='Excel workbook. When grouping is on, each length class '
                             'goes to its own sheet.')
                else:
                    e3.button('XLSX - openpyxl missing', disabled=True,
                              use_container_width=True,
                              help='Zainstaluj: pip install openpyxl')

                st.caption('Files contain the candidates shown after '
                           'filtering. To export everything, select all '
                           'off-target statuses above.')

                st.session_state['widoczne'] = widoczne


# --- 4. KONSTRUKTY -----------------------------------------------------------

with zakl[4]:
    if 'ranking' not in st.session_state:
        st.info('Generate candidates first.')
    else:
        rank = st.session_state['ranking']

        naglowek_sekcji('plazmid', 'Expression system')
        st.caption('Select a promoter and a cloning method. The promoter '
                   'determines whether a start nucleotide and terminator must '
                   'be added; the cloning method determines which flanking '
                   'sequences are appended.')

        # --- rejestr wlasnych elementow uzytkownika --------------------------
        if 'wlasne_promotory' not in st.session_state:
            st.session_state['wlasne_promotory'] = {}
        if 'wlasne_klonowanie' not in st.session_state:
            st.session_state['wlasne_klonowanie'] = {}

        dostepne_prom = {**constructs.PROMOTORY,
                         **st.session_state['wlasne_promotory']}
        dostepne_klon = {**constructs.KLONOWANIE,
                         **st.session_state['wlasne_klonowanie']}

        with st.expander('Add a custom promoter or vector'):
            st.caption(
                'The registry holds the requirements of common systems but '
                '**not the sequences of specific vectors**. pCAMBIA variants '
                'differ between laboratories, so a sequence entered from '
                'memory would give a construct that does not match the '
                'plasmid in your freezer. Paste sequences from your own '
                'vector map.')

            t1, t2 = st.tabs(['Nowy promotor', 'Nowa metoda klonowania'])

            with t1:
                np_nazwa = st.text_input('Nazwa', key='np_n',
                                         placeholder='np. AtU6-26')
                np_pol = st.selectbox('Polimeraza', ['PolIII', 'PolII'],
                                      key='np_p')
                np_nt = st.selectbox(
                    'Wymagany pierwszy nukleotyd transkryptu',
                    ['brak wymogu', 'G', 'A', 'C', 'T'], key='np_nt',
                    help='Pol III wymusza pierwszy nukleotyd: U6 wymaga G, '
                         'U3 wymaga A. Pol II nie ma takiego wymogu.')
                np_term = st.text_input(
                    'Terminator', key='np_t',
                    value='TTTTTT' if np_pol == 'PolIII' else 'NOS-ter (element osobny)',
                    help='For Pol III the T-tract goes into the insert. For Pol II '
                         'terminator jest osobnym elementem wektora.')
                np_max = st.number_input('Max. transcript length (nt)',
                                         0, 5000,
                                         250 if np_pol == 'PolIII' else 0,
                                         key='np_m',
                                         help='0 oznacza brak limitu.')
                np_uw = st.text_area('Uwagi', key='np_u', height=68)

                if st.button('Dodaj promotor', key='np_btn'):
                    if not np_nazwa.strip():
                        st.error('Enter a name.')
                    else:
                        st.session_state['wlasne_promotory'][np_nazwa] = \
                            constructs.Promotor(
                                nazwa=np_nazwa, polimeraza=np_pol,
                                wymagany_pierwszy_nt=(
                                    None if np_nt == 'brak wymogu' else np_nt),
                                terminator=np_term,
                                max_dlugosc_nt=np_max if np_max > 0 else None,
                                uwagi=np_uw or 'Element zdefiniowany przez '
                                               'user.')
                        st.success(f'Dodano: {np_nazwa}')
                        st.rerun()

            with t2:
                nk_nazwa = st.text_input('Nazwa', key='nk_n',
                                         placeholder='np. pCAMBIA1300 / XbaI-SacI')
                st.caption('Paste flanking sequences from your own vector map. '
                           'They are appended on both sides of the insert.')
                nk_f5 = st.text_input("Sekwencja przed insertem (5′)",
                                      key='nk_5',
                                      placeholder='AAAATCTAGA')
                nk_f3 = st.text_input("Sekwencja za insertem (3′)",
                                      key='nk_3',
                                      placeholder='GAGCTCAAAA')
                nk_zabr = st.text_input(
                    'Motywy zabronione w insercie (oddziel przecinkiem)',
                    key='nk_z', placeholder='TCTAGA, GAGCTC',
                    help='Recognition sites of the enzymes used. If present inside '
                         'the insert, the enzyme will cut the construct in '
                         'the wrong place.')
                nk_uw = st.text_area('Uwagi', key='nk_u', height=68)

                if st.button('Add method', key='nk_btn'):
                    if not nk_nazwa.strip():
                        st.error('Enter a name.')
                    else:
                        motywy = [m.strip().upper()
                                  for m in nk_zabr.split(',') if m.strip()]
                        st.session_state['wlasne_klonowanie'][nk_nazwa] = \
                            constructs.MetodaKlonowania(
                                nazwa=nk_nazwa,
                                flanka_5=nk_f5.strip().upper(),
                                flanka_3=nk_f3.strip().upper(),
                                zabronione_motywy=motywy,
                                zostaje_w_konstrukcie='wg definicji '
                                                      'user',
                                uwagi=nk_uw or 'Element zdefiniowany przez '
                                               'user.')
                        st.success(f'Dodano: {nk_nazwa}')
                        st.rerun()

            if (st.session_state['wlasne_promotory']
                    or st.session_state['wlasne_klonowanie']):
                st.divider()
                st.caption('Zdefiniowane w tej sesji:')
                for n in st.session_state['wlasne_promotory']:
                    st.markdown(f'<span class="plakietka p-roz">promotor: '
                                f'{n}</span>', unsafe_allow_html=True)
                for n in st.session_state['wlasne_klonowanie']:
                    st.markdown(f'<span class="plakietka p-srebro">'
                                f'klonowanie: {n}</span>',
                                unsafe_allow_html=True)
                st.caption('Definitions are lost when the app closes. To keep '
                           'them permanently, add them to the registry '
                           'w pliku constructs.py.')

        c1, c2, c3 = st.columns(3)
        prom = c1.selectbox('Promoter', sorted(dostepne_prom))
        klon = c2.selectbox('Cloning method', sorted(dostepne_klon))
        pet = c3.selectbox('shRNA loop', sorted(constructs.PETLE))

        p, k = dostepne_prom[prom], dostepne_klon[klon]
        st.markdown(
            f'<span class="plakietka p-fiolet">{p.polimeraza}</span>'
            f'<span class="plakietka p-zloto">first nt: '
            f'{p.wymagany_pierwszy_nt or "dowolny"}</span>',
            unsafe_allow_html=True)
        with st.expander('System requirements'):
            st.caption(p.uwagi)
            st.caption(k.uwagi)

        st.divider()
        st.markdown('#### Candidate selection')
        zrodlo_c = st.session_state.get('widoczne') or rank
        nazwy_c = [c['nazwa'] for c in zrodlo_c]

        sposob = st.radio(
            'Selection', ['Top N by rank', 'Pick individually'],
            horizontal=True, label_visibility='collapsed')

        if sposob == 'Top N by rank':
            n = st.number_input('Number of constructs', 1,
                                max(len(nazwy_c), 1),
                                min(3, len(nazwy_c)))
            st.caption(f'{len(nazwy_c)} candidates available after filtering. '
                       f'Pareto-front candidates are taken first.')
            wybrane_c = None
        else:
            wybrane_c = st.multiselect(
                'Candidates to generate', nazwy_c,
                default=nazwy_c[:3],
                help='Pick exactly which candidates to build. Useful when '
                     'the highest-ranked one is not the one you want to '
                     'order.')
            n = len(wybrane_c)
        if wybrane_c is not None:
            wyb = [c for c in zrodlo_c if c['nazwa'] in wybrane_c]
        else:
            wyb = ([c for c in zrodlo_c if c.get('pareto')][:n]
                   or zrodlo_c[:n])

        for c in wyb:
            kas = zbuduj_z_rejestru(c['nazwa'], c['guide_rna'],
                                    p, k, pet)
            with st.container(border=True):
                st.markdown(f'#### {kas.nazwa}')
                a, b = st.columns(2)
                a.caption('Guide strand (antisense)')
                a.code(f"5'-{kas.guide}-3'")
                b.caption('Passenger strand (sense)')
                b.code(f"5'-{kas.passenger}-3'")
                st.caption(f'shRNA — {len(kas.shrna_dna)} nt')
                st.code(kas.shrna_dna)
                st.caption(f'Cassette for ordering - {len(kas.kaseta_dna)} nt')
                st.code(kas.kaseta_dna)
                for o in kas.ostrzezenia:
                    st.warning(o)


# --- 5. METODYKA -------------------------------------------------------------

with zakl[5]:
    naglowek_sekcji('helisa', 'Laboratory order sheet')
    st.caption('A ranked guide sequence is not something that can be ordered '
               'and used at the bench. Depending on the experimental route, '
               'several distinct constructs are needed, each with its own end '
               'chemistry and flanking sequence.')

    if not OLIGOS_OK:
        st.error('Module oligos.py unavailable.')
    elif 'ranking' not in st.session_state:
        st.info('Generate candidates first.')
    else:
        zrodlo_list = st.session_state.get('widoczne') or st.session_state['ranking']
        nazwy = [c['nazwa'] for c in zrodlo_list]
        sel1, sel2 = st.columns([3, 1])
        wybrane = sel1.multiselect('Candidates', nazwy, default=nazwy[:3])
        if sel2.button('Select all', use_container_width=True):
            wybrane = nazwy
        sel2.caption(f'{len(nazwy)} available')

        o1, o2 = st.columns(2)
        formy = o1.multiselect(
            'Forms to generate',
            ['duplex', 'shrna', 'cloning', 'ivt', 'detection', 'controls'],
            default=['duplex', 'shrna', 'controls'],
            help='duplex — synthetic strands with 3-prime overhangs · '
                 'shrna — hairpin · cloning — annealing oligos · '
                 'ivt — T7 templates · detection — stem-loop RT-qPCR and '
                 'Northern probe · controls — C911, seed mutant, scrambled')
        overhang = o2.selectbox(
            '3-prime overhang', ['dTdT', 'UU', 'native'],
            help='dTdT is conventional: cheaper and more nuclease-resistant '
                 '(Elbashir et al. 2001).')
        petla_o = st.selectbox('shRNA loop', sorted(oligos.LOOPS),
                               format_func=lambda k: f'{k} — {oligos.LOOPS[k][1]}')

        if wybrane:
            arkusze = []
            wiersze_all = []
            for nazwa in wybrane:
                kand = next(c for c in zrodlo_list if c['nazwa'] == nazwa)
                sh = oligos.build_order_sheet(
                    nazwa, kand['guide_rna'],
                    target_mrna=st.session_state.get('mrna'),
                    loop_key=petla_o, overhang=overhang,
                    include=tuple(formy))
                arkusze.append(sh)
                wiersze_all.extend(oligos.order_sheet_to_rows(sh))

            for sh in arkusze:
                with st.container(border=True):
                    st.markdown(f'#### {sh.name}')
                    if 'duplex' in sh.forms:
                        st.caption('Annealed duplex')
                        st.code(sh.forms['duplex']['diagram'])
                    if 'controls' in sh.forms:
                        c = sh.forms['controls']
                        st.caption('Specificity controls')
                        cc1, cc2 = st.columns(2)
                        cc1.markdown('**C911 — recommended**')
                        cc1.code(c['c911_5_3'])
                        cc1.caption(c['c911_note'])
                        cc2.markdown('**Scrambled — weak**')
                        cc2.code(c['scrambled_5_3'])
                        cc2.caption(c['scrambled_note'])
                    for w in sh.warnings:
                        st.warning(w)

            st.divider()
            st.markdown('#### Full order sheet')
            kol_o = ['candidate', 'category', 'form', 'chemistry',
                     'sequence_5_3', 'length_nt', 'gc_percent']
            st.dataframe(wiersze_all, use_container_width=True, height=380)

            x1, x2 = st.columns(2)
            x1.download_button('CSV', do_tsv(wiersze_all, kol_o, sep=';'),
                               'siRNA_order_sheet.csv', 'text/csv',
                               use_container_width=True)
            xl = do_xlsx({'order_sheet': wiersze_all}, kol_o)
            if xl:
                x2.download_button(
                    'XLSX', xl, 'siRNA_order_sheet.xlsx',
                    'application/vnd.openxmlformats-officedocument.'
                    'spreadsheetml.sheet', use_container_width=True)

            # ---------- IN SILICO SYNTHESIS REPORT ----------
            if REPORT_OK:
                st.divider()
                st.markdown('#### In silico synthesis report')
                st.caption('A complete, self-contained record of the design: '
                           'inputs with a checksum, every parameter including '
                           'defaults, candidate metrics, sequences to order, '
                           'and the limitations of the analysis. A design '
                           'that cannot be reproduced from its report is not '
                           'a result.')

                ctx = report.ReportContext(
                    target_name=st.session_state.get('naglowek', 'target'),
                    target_sequence=st.session_state.get('mrna', ''),
                    host_profile=f'{profil.nazwa} — {profil.opis}',
                    lengths=tuple(sorted(dlugosci)),
                    parameters={
                        'gc_min': gc_min, 'gc_max': gc_max,
                        'min_asymmetry': min_asym,
                        'max_guide_mfe': max_mfe,
                        'exclude_5prime_nt': w5,
                        'exclude_3prime_nt': w3,
                        'advanced_mode': zaawansowany,
                        'forbidden_motifs': (', '.join(profil.motywy_zabronione)
                                             or 'none'),
                        'guide_5prime_required': (profil.guide_5_dozwolone
                                                  or 'any'),
                        'shRNA_loop': petla_o,
                        'duplex_overhang': overhang,
                    },
                    transcriptome_file=(
                        'uploaded' if 'sciezka_tx' in st.session_state
                        else None),
                    transcriptome_size_mb=(
                        offtarget.szacuj_czas(
                            st.session_state['sciezka_tx'])['rozmiar_mb']
                        if 'sciezka_tx' in st.session_state and OFFTARGET_OK
                        else None),
                    isolate_file=('uploaded'
                                  if 'sciezka_izolaty' in st.session_state
                                  else None),
                    n_isolates=(st.session_state['kons']['n_sekwencji']
                                if 'kons' in st.session_state else None),
                    conservation_threshold=(
                        st.session_state['opt']['threshold']
                        if 'opt' in st.session_state else None),
                    conservation_method=(
                        st.session_state['opt']['method']
                        if 'opt' in st.session_state else None),
                    tool_versions={
                        'ViennaRNA': ('active' if design.VIENNA_DOSTEPNE
                                      else 'unavailable'),
                    })

                wybrani = [c for c in zrodlo_list if c['nazwa'] in wybrane]
                md = report.build_markdown(
                    ctx, wybrani,
                    sensitivity=st.session_state.get('wrazliwosc'),
                    order_rows=wiersze_all,
                    n_filtered=st.session_state.get('odrzucone'))

                st.markdown(
                    f'<span class="plakietka p-fiolet">'
                    f'{len(wybrani)} candidates in report</span>'
                    + ''.join(f'<span class="plakietka p-srebro">{c["nazwa"]}'
                              f'</span>' for c in wybrani),
                    unsafe_allow_html=True)
                st.caption('The report contains exactly the candidates '
                           'selected above, with their full metrics and the '
                           'sequences to order. Change the selection to '
                           'change the report.')

                r1, r2, r3 = st.columns(3)
                r1.download_button(
                    'Markdown', md, 'siRNA_design_report.md',
                    'text/markdown', use_container_width=True)
                r2.download_button(
                    'HTML (printable)',
                    report.build_html(md, 'siRNA design report'),
                    'siRNA_design_report.html', 'text/html',
                    use_container_width=True,
                    help='Open in a browser and print with Ctrl+P; choose '
                         '"Save as PDF" for a PDF file.')
                dx = report.build_docx(
                    ctx, wybrani,
                    sensitivity=st.session_state.get('wrazliwosc'),
                    order_rows=wiersze_all,
                    n_filtered=st.session_state.get('odrzucone'))
                if dx:
                    r3.download_button(
                        'Word (.docx)', dx, 'siRNA_design_report.docx',
                        'application/vnd.openxmlformats-officedocument.'
                        'wordprocessingml.document', use_container_width=True)
                else:
                    r3.button('Word - python-docx missing', disabled=True,
                              use_container_width=True,
                              help='Install with: pip install python-docx')

                with st.expander('Preview'):
                    st.markdown(md)


with zakl[6]:
    naglowek_sekcji('ksiazka', 'Methods')

    st.markdown("""
The computational basis of each stage is described below, with references to
source publications. Numbering corresponds to the reference list at the end.
""")

    with st.expander('1 · Duplex thermodynamics'):
        st.markdown("""
The free energy of the RNA–RNA duplex is computed by the nearest-neighbour
method using the parameters determined by **Xia et al. (1998)** [1] for
Watson–Crick pairs in RNA. The model sums the contributions of all adjacent
base pairs and adds a helix-initiation term together with a penalty for
terminal A–U pairs.

These parameters are the standard in the field and are identical to the
stacking component used in the ViennaRNA package [2], which allows direct
comparison of results.

#### Thermodynamic asymmetry

The key quantity derived from this model is **asymmetry** — the difference in
free energy of the four base pairs at the 5-prime end of the guide strand and
of the passenger strand:

```
asymmetry = dG5'(guide) − dG5'(passenger)
```

The RISC complex unwinds the duplex and retains only the strand whose 5-prime
end is **more weakly** bound, because unwinding starts more readily from that
end. This relationship was described independently by **Khvorova et al.
(2003)** [3] and **Schwarz et al. (2003)** [4].

A positive value means RISC will load the intended strand. A negative value
means the passenger strand is loaded — the siRNA will not act on its target
and may silence unintended transcripts.

**Limitation.** The parameters of Xia et al. were determined in 1 M NaCl. The
reported melting temperature carries no ionic-strength correction and does not
correspond to physiological conditions. For ranking candidates against one
another this is immaterial, since all are affected equally.
""")

    with st.expander('2 · Secondary structure and target accessibility'):
        st.markdown("""
Secondary structures are folded using **ViennaRNA** (**Lorenz et al. 2011**)
[2] as the reference implementation of the Turner thermodynamic model.

Two quantities are computed.

**Guide-strand MFE** — a strand with strong intramolecular structure is
occupied by self-interaction and will not bind its target. Candidates with a
minimum free energy below the threshold are rejected.

**Target-site accessibility** — the free energy of a local mRNA window
spanning the target site and its surroundings, normalised per nucleotide. A
strongly structured context impedes access by the RISC complex. The importance
of target-site accessibility for silencing efficiency was shown by
**Shabalina et al. (2006)** [5] and **Matveeva et al. (2007)** [6].

**Limitation.** A window free-energy approximation is used. A more accurate
approach relies on unpaired probability derived from the Boltzmann ensemble
(`RNAplfold`); the corresponding function is implemented but disabled by
default on grounds of computational cost.
""")

    with st.expander('3 · Physicochemical filters'):
        st.markdown("""
The rule set draws on the work that first systematised the relationship
between sequence features and silencing efficiency: **Elbashir et al. (2001)**
[7], **Reynolds et al. (2004)** [8] and **Ui-Tei et al. (2004)** [9].

| criterion | rationale |
|---|---|
| GC content within the profile range | below the lower bound the duplex is too weak; above the upper bound RISC will not unwind it |
| no runs of ≥ 4 identical nucleotides | impede chemical synthesis; G-runs favour G-quadruplex formation |
| 5-prime end of the guide strand | AGO preference, host-dependent — see below |
| 5-prime end of the passenger strand | reinforces asymmetry in the desired direction |
| exclusion of ORF termini | regions shielded by the translation initiation complex and by termination factors |

#### Host dependence

Design rules are **not universal across kingdoms**, which is why the tool
applies organism profiles.

**Plants.** AGO1 shows MID-pocket selectivity for 5-prime U, AGO2 for 5-prime
A and AGO5 for 5-prime C, as shown by **Mi et al. (2008)** [10] and
**Takeda et al. (2008)** [11]. Because the antiviral pathway runs mainly
through AGO1 and AGO2, the plant profile requires 5-prime U or A.

**Mammals.** Immunostimulatory motifs recognised by TLR7 and TLR8 are
filtered: `UGUGU` (**Judge et al. 2005**) [12] and `GUCCUUCAA`
(**Hornung et al. 2005**) [13]. In the plant profile this filter is disabled,
since plants possess neither Toll-like receptors nor an interferon pathway.
The motifs are nonetheless reported.

Plants do, however, mount a sequence-independent response to double-stranded
RNA (dsRNA-PTI), which cannot be eliminated at the design stage and requires
an experimental control in the form of dsRNA of neutral sequence.
""")

    with st.expander('4 · Off-target analysis'):
        st.markdown("""
#### Why not BLAST

The **BLAST** algorithm (**Altschul et al. 1990**) [14] was designed to find
homology between long sequences. For queries of 21 nt two of its mechanisms
work against us:

- the default `word_size` of 11 means that shorter matches are never seeded —
  **a seed region of 7–8 nt is invisible by construction**;
- the E-value depends on query length and remains high for 21 nt even for a
  perfect match, so the significance filter discards biologically meaningful
  hits.

BLAST also treats all positions as equivalent, whereas in an siRNA positions
2–8 and the cleavage site matter incomparably more than the 3-prime region.

#### The approach used here

Streaming scan with indexing of the **queries** rather than the transcriptome.
Memory use is proportional to the number of candidates, not to the size of the
transcriptome.

**Seed-region hits** are classified according to the scheme of
**Bartel (2009)** [15], standard in miRNA biology:

| class | definition |
|---|---|
| 8mer | complementarity of positions 2–8 plus an A opposite position 1 — the strongest class |
| 7mer-m8 | complementarity of positions 2–8 |
| 7mer-A1 | complementarity of positions 2–7 plus an A opposite position 1 |

**Full-length hits** are scored with positional weights:

| guide position | weight | rationale |
|---|---|---|
| 1 | 0.5 | does not pair with the target; housed in the MID pocket of AGO |
| 2–8 | 3.0 | seed region, determines target recognition |
| 9–11 | 2.0 | cleavage site; a mismatch blocks endonucleolytic activity |
| 12 onwards | 1.0 | 3-prime region, supports binding but does not determine recognition |

#### Correction for database size

An exact 8-nucleotide match occurs by chance once every 4⁸ = 65 536
positions, that is roughly 1800 times in a database of 1.2 × 10⁸ nucleotides.
A criterion of the form "any 8mer hit rejects the candidate" is therefore
worthless at that scale. Equally, a minimum cost taken over tens of thousands
of random positions is necessarily low — this is a multiple-testing problem.

The criterion used instead is the **critical hit**: at most three mismatches
and none in the seed region. The expected number of such matches in a database
of 10⁸ nucleotides is negligible, so their occurrence is a real signal. The
count of 8mer hits is reported as **enrichment relative to chance
expectation**, not as an absolute value.

#### Recognition of the intended target

If the target gene is present in the host transcriptome — which is always the
case when targeting an endogenous gene — every candidate hits its own target
at zero cost. Without distinguishing intended from unintended hits, all
candidates would be rejected.

The solution: for each hit the surrounding context is checked against the
target mRNA. If it is found there, the hit is classified as intended and
excluded from the off-target assessment.
""")

    with st.expander('5 · Conservation'):
        st.markdown("""
When designing an siRNA against a pathogen, the target is a moving one — a
viral population is a set of isolates differing in sequence, and a mutation at
the target site allows escape from silencing.

Instead of the classical Shannon entropy derived from a multiple sequence
alignment, **exact k-mer coverage** is used: for each window in the reference
sequence, the fraction of remaining isolates containing an identical string is
computed.

The rationale is biological. An siRNA requires near-perfect complementarity,
and a single mismatch in the seed region is enough for the target not to be
recognised [15]. The relevant quantity is therefore not conservation in the
evolutionary sense but the presence of a specific k-mer in a given isolate —
which is measured directly, without alignment.

The strategy of targeting conserved regions, and of targeting several genomic
sites simultaneously to counter escape, was described by
**ter Brake et al. (2006)** [16].

**Limitation.** The method detects substitutions only. Insertions and
deletions would require a multiple sequence alignment; Shannon entropy is
implemented in the module and available for pre-aligned input.
""")

    with st.expander('6 · Scoring and ranking'):
        st.markdown("""
#### Normalisation

Every metric is min–max normalised to the interval [0, 1] **before** its
weight is applied. This step is necessary: metrics with different numerical
ranges (asymmetry from −6 to +6 kcal/mol, GC content from 30 to 64 per cent)
would otherwise contribute in proportion to their own range rather than to the
declared weight.

#### Sensitivity analysis

The weights are literature-informed but remain an arbitrary choice.
Verification consists of altering each weight by ±25 per cent and comparing
the top five of the ranking. Stability close to 1.0 supports the claim that
the ordering of candidates does not depend on the choice of weights.

#### Pareto front

The set of non-dominated candidates is determined — those that cannot be
improved in one criterion without worsening another. A candidate scored highly
but lying **outside** the front implies the existence of a candidate better in
every criterion simultaneously; its position then derives from the weighting,
not from properties of the sequence.

#### Direction of development

The intended solution is to replace manual weights with a model trained on
measured efficacy data. The reference set remains that of
**Huesken et al. (2005)** [17], comprising 2431 siRNA molecules of measured
activity; composite sets combining data from several independent studies are
also available.

**Caveat.** All available training sets derive from mammalian systems. A model
trained on them and applied to plants would carry an error requiring explicit
discussion. A plant set of comparable size does not exist.
""")

    with st.expander('7 · Expression constructs and laboratory forms'):
        st.markdown("""
The shRNA architecture comprises guide strand, loop and passenger strand, with
the guide placed at the 5-prime end to favour its loading into the effector
complex. The design and validation of shRNA constructs are discussed by
**Taxman et al. (2006)** [18].

Requirements depend on two independent factors.

**Polymerase.** Pol III promoters (U6, U3) require a defined first nucleotide
of the transcript and a poly-T terminator placed inside the insert; the
transcript receives neither a cap nor a poly-A tail, which favours a hairpin
with defined ends. Pol II promoters (35S, UBQ10) impose no requirement on the
first nucleotide, but the terminator must be a separate element of the vector.

**Cloning method.** Restriction cloning, Golden Gate, Gateway and Gibson each
leave different flanking sequences in the finished construct, which for a
short hairpin may affect its folding.

The registry holds the requirements of common systems but **not the sequences
of specific vectors** — pCAMBIA variants and their derivatives differ between
laboratories, so flanking sequences must be entered from the map of the vector
in hand.

#### Specificity controls

| control | construction | property |
|---|---|---|
| **C911** | positions 9–11 replaced by their complement | **recommended** — retains the seed and hence off-target activity, while central mismatches abolish target cleavage |
| C10 | position 10 only | weaker variant of the above |
| seed mutant | positions 2–8 replaced by their complement | mirror image of C911 — abolishes seed-mediated off-target activity |
| scrambled | shuffled composition | conventional but weak |

**Buehler et al. (2012)** [19] compared these designs on 20 highly active
siRNAs comprising 10 true and 10 false positives. Scrambled controls lost
activity in both groups and could not distinguish them. C911 separated the two
groups completely: false positives retained most of their activity, true
positives lost it.

A difference between the parent siRNA and its C911 control is therefore
attributable to on-target activity.
""")

    st.divider()
    st.markdown('### Organism profiles')
    st.code(hosts.lista_profili())

    st.divider()
    st.markdown("""
### Limitations

1. The off-target scan runs in pure Python at roughly 0.5 million nucleotides
   per second. For transcriptomes of human genome size an FM-index would be
   appropriate.
2. Target accessibility is estimated from window free energy rather than
   unpaired probability.
3. Scoring weights are set manually; no trained model is used.
4. Melting temperature carries no ionic-strength correction.
5. Conservation analysis detects substitutions only.
6. No validation against measured efficacy data — planned using set [17].

### References

1. Xia T., SantaLucia J., Burkard M.E. et al. (1998) Thermodynamic parameters
   for an expanded nearest-neighbor model for formation of RNA duplexes with
   Watson-Crick base pairs. *Biochemistry* 37:14719–14735.
2. Lorenz R., Bernhart S.H., Höner zu Siederdissen C. et al. (2011)
   ViennaRNA Package 2.0. *Algorithms for Molecular Biology* 6:26.
3. Khvorova A., Reynolds A., Jayasena S.D. (2003) Functional siRNAs and
   miRNAs exhibit strand bias. *Cell* 115:209–216.
4. Schwarz D.S., Hutvágner G., Du T. et al. (2003) Asymmetry in the assembly
   of the RNAi enzyme complex. *Cell* 115:199–208.
5. Shabalina S.A., Spiridonov A.N., Ogurtsov A.Y. (2006) Computational models
   with thermodynamic and composition features improve siRNA design.
   *BMC Bioinformatics* 7:65.
6. Matveeva O., Nechipurenko Y., Rossi L. et al. (2007) Comparison of
   approaches for rational siRNA design leading to a new efficient and
   transparent method. *Nucleic Acids Research* 35:e63.
7. Elbashir S.M., Harborth J., Lendeckel W. et al. (2001) Duplexes of
   21-nucleotide RNAs mediate RNA interference in cultured mammalian cells.
   *Nature* 411:494–498.
8. Reynolds A., Leake D., Boese Q. et al. (2004) Rational siRNA design for
   RNA interference. *Nature Biotechnology* 22:326–330.
9. Ui-Tei K., Naito Y., Takahashi F. et al. (2004) Guidelines for the
   selection of highly effective siRNA sequences for mammalian and chick
   RNA interference. *Nucleic Acids Research* 32:936–948.
10. Mi S., Cai T., Hu Y. et al. (2008) Sorting of small RNAs into Arabidopsis
    argonaute complexes is directed by the 5′ terminal nucleotide.
    *Cell* 133:116–127.
11. Takeda A., Iwasaki S., Watanabe T. et al. (2008) The mechanism selecting
    the guide strand from small RNA duplexes is different among Argonaute
    proteins. *Plant and Cell Physiology* 49:493–500.
12. Judge A.D., Sood V., Shaw J.R. et al. (2005) Sequence-dependent
    stimulation of the mammalian innate immune response by synthetic siRNA.
    *Nature Biotechnology* 23:457–462.
13. Hornung V., Guenthner-Biller M., Bourquin C. et al. (2005)
    Sequence-specific potent induction of IFN-α by short interfering RNA in
    plasmacytoid dendritic cells through TLR7. *Nature Medicine* 11:263–270.
14. Altschul S.F., Gish W., Miller W. et al. (1990) Basic local alignment
    search tool. *Journal of Molecular Biology* 215:403–410.
15. Bartel D.P. (2009) MicroRNAs: target recognition and regulatory
    functions. *Cell* 136:215–233.
16. ter Brake O., Konstantinova P., Ceylan M., Berkhout B. (2006) Silencing
    of HIV-1 with RNA interference: a multiple shRNA approach.
    *Molecular Therapy* 14:883–892.
17. Huesken D., Lange J., Mickanin C. et al. (2005) Design of a genome-wide
    siRNA library using an artificial neural network.
    *Nature Biotechnology* 23:995–1001.
18. Taxman D.J., Livingstone L.R., Zhang J. et al. (2006) Criteria for
    effective design, construction, and gene knockdown by shRNA vectors.
    *BMC Biotechnology* 6:7.
19. Buehler E., Chen Y.-C., Martin S. (2012) C911: a bench-level control for
    sequence specific siRNA off-target effects. *PLoS ONE* 7(12):e51942.
    DOI 10.1371/journal.pone.0051942.
20. Varkonyi-Gasic E., Wu R., Wood M. et al. (2007) Protocol: a highly
    sensitive RT-PCR method for detection and quantification of microRNAs.
    *Plant Methods* 3:12.

*Bibliographic details should be verified before citation in a publication.*
""")


st.markdown('<div class="stopka">siRNA Design Studio · Antonina Jarecka</div>',
            unsafe_allow_html=True)
